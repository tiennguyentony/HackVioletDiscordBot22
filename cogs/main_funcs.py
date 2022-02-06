from discord.ext import commands
import openpyxl
import random

wb = openpyxl.load_workbook("Hackathon.xlsx")
sh = wb['Hackathons']
w_b = wb.active
wb.active = 1
sh2 = wb['Opportunities']
w_a = wb.active


class main_funcs(commands.Cog):

    def __init__(self,bot):
        self.bot = bot
    @commands.command()
    async def encourage(self,ctx):
        encourage_messages = ["You are more powerful than you know!","Believe you can and you're halfway there!", "There's something in you that the world needs!","You are so much stronger than you think!", "Every day may not be good, but there's something good in every day!", "The best time for new beginnings is NOW!","Nothing is impossible!", "When you feel like quitting, think about why you started","Don't Give Up!", "Everything happens for a reason, a season or for a lifetime"]
        response = encourage_messages[random.randint(0, len(encourage_messages)-1)]
        await ctx.send(response)
#displays a list of upcoming hackathons
    @commands.command()
    async def hackathons(self, ctx, para = None):
        print(wb.sheetnames)
        response = ""

        for row in range(2, (len(list(w_b.rows))+1)):

            if para.lower() == sh.cell(row, 4).value.lower():
                hack = f"\n **{sh.cell(row, 2).value}** is an {sh.cell(row, 6).value.lower()} hackathon, hosted by **{sh.cell(row, 3).value}** will take place on _{sh.cell(row, 5).value} {sh.cell(row, 4).value}_."
                hacklink = f"\n    {sh.cell(row,1).value}"

                if (len(response)+len(hack)+len(hacklink))<=2000:
                    response += (hack + hacklink)
                    response += "\n"

        for row in range(2, (len(list(w_b.rows))+1)):

            if para.lower() == sh.cell(row, 6).value.lower():
                hack = f"\n**{sh.cell(row, 2).value}** is an {sh.cell(row, 6).value.lower()} hackathon, hosted by **{sh.cell(row, 3).value}** will take place on _{sh.cell(row, 5).value} {sh.cell(row, 4).value}_."
                hacklink =  f"\n   {sh.cell(row,1).value}"

                if (len(response)+len(hack)+len(hacklink))<=2000:
                    response += (hack + hacklink) 
                    response += "\n"

        if para == "all":

            for row in range(2, (len(list(w_b.rows))+1)):
                hack = f"\n**{sh.cell(row, 2).value}** is an {sh.cell(row, 6).value.lower()} hackathon, hosted by **{sh.cell(row, 3).value}** will take place on _{sh.cell(row, 5).value} {sh.cell(row, 4).value}_."
                hacklink =  f"\n    {sh.cell(row,1).value}"

                if (len(response)+len(hack)+len(hacklink))<=2000:
                    response += (hack + hacklink)
                    response += "\n"

        await ctx.send(response)
    
   
#displays a list of sites to search for opportunities in Computer Science
    @commands.command()
    async def opportunities(self, ctx):
        response = ""

        for row in range(2, (len(list(w_a.rows)) + 1)):

            for col in reversed(range(1, 3)):

                if(col ==1):
                    response += "       "
                    response += sh2.cell(row, col).value

                if(col ==2):
                    response += f"**{sh2.cell(row, col).value}**"
                    response += ": "

                response += "\n"
                
        await ctx.send(response)
    @commands.command()
    async def python(self, ctx):
        response = ("""
        Here is some resources that I found: 

        Learning Resrouces you might want to start: 
        https://www.learnpython.org/

        Python class: 
        https://developers.google.com/edu/python

        Basic Usage of Python: 
        https://docs.python-guide.org/

        Project Ideas:
            - Discord Bot 
            - Microblog
            - Quiz Application """)

        await ctx.send(response)
    @commands.command()
    async def Cplusplus (self, ctx):
        response = ("""
        Here is some resources that I found: 

        Learning Resrouces you might want to start: 
        https://www.learncpp.com/

        C++ class: 
        https://developers.google.com/edu/c++

        What C++ used for: 
        https://www.softwaretestinghelp.com/cpp-applications/

        Project Ideas:
            - Login and Registeration System 
            - Sudoku Game
            - Car Rental System """)

        await ctx.send(response)
    @commands.command()
    async def Typescript (self, ctx):
        response = ("""
        Here is some resources that I found: 

        Learning Resrouces you might want to start: 
        https://www.typescriptlang.org/docs/
        Follow this for Typescript weekly information: 
        https://www.typescript-weekly.com/

        What Typescript used for: 
        https://javascript.plainenglish.io/7-trending-typescript-projects-on-github-675d3fc8ecae

        Project Ideas:
            - Restaunrant Menu
            - Navigation Bar on scroll
            - Countdown clock""")

        await ctx.send(response)
    @commands.command()
    async def Ruby (self, ctx):
        response = ("""
        Here is some resources that I found: 

        Learning Resrouces you might want to start: 
        https://guides.rubyonrails.org/getting_started.html
        
        Valuable book to learn Ruby:
        https://www.railstutorial.org/

        What makes Ruby different from other: 
        https://www.codecademy.com/resources/blog/what-is-ruby-used-for/

        Project Ideas:
            - Make a website with Hanami
            - Develop a Chat Widget using Chatwoot 
            - Build a 2D game with Goby """)

        await ctx.send(response)
    @commands.command()
    async def Ada (self, ctx):
        response = ("""
        Here is some resources that I found: 

        Learning Resrouces you might want to start: 
        https://www.adaic.org/learn/
        
        Valuable pdfs to learn Ada:
        https://learn.adacore.com/

        What makes Ada different from other: 
        https://people.cs.kuleuven.be/~dirk.craeynest/ada-belgium/ada.html

        Project Ideas:
            - Payroll systems
            - Commercial banking systems
            - Stock quotation transaction systems
            - Language translation system and RDMSs """)

        await ctx.send(response)
    @commands.command()
    async def Cobol (self, ctx):
        response = ("""
        Here is some resources that I found: 

        Learning Resrouces you might want to start: 
        https://www.guru99.com/learn-cobol-programming-tutorial.html
        
        Another tutorial for Cobol:
        https://www.tutorialspoint.com/cobol/index.htm

        What makes Ada different from other: 
        https://searchitoperations.techtarget.com/definition/COBOL-Common-Business-Oriented-Language

        Project Ideas:
            - Read a record
            - Process a record
            - Write a record """)

        await ctx.send(response)
    @commands.command()
    async def PowerShell (self, ctx):
        response = ("""
        Here is some resources that I found: 

        Learning Resrouces you might want to start: 
        https://www.varonis.com/blog/windows-powershell-tutorials#:~:text=What%20is%20PowerShell%20Language%3F,be%20used%20in%20Windows%20environments.
        
        Another tutorial for Powershell:
        https://docs.microsoft.com/en-us/powershell/scripting/overview?view=powershell-7.2

        How can you do the Powershell: 
        https://ccbtechnology.com/what-is-powershell/

        Project Ideas:
            - Firewall status
            - Current CPU/RAM/Disk usage
            - Network info""")

        await ctx.send(response)
    @commands.command()
    async def Java (self, ctx):
        response = ("""
        Here is some resources that I found: 

        Learning Resrouces you might want to start: 
        https://www.w3schools.com/java/default.asp  

        Another tutorial for Java:
        https://www.tutorialspoint.com/java/index.htm

        Why Java is important:
        https://codeinstitute.net/global/blog/what-is-java/

        Project Ideas:
            - Airline Reservation System
            - Data Visualization Software
            - Electricity Billing System""")

        await ctx.send(response)
    @commands.command()
    async def PHP (self, ctx):
        response = ("""
        Here is some resources that I found: 

        Learning Resrouces you might want to start: 
        https://www.w3schools.com/php/  

        Another tutorial for PHP:
        https://www.tutorialspoint.com/php/index.htm

        Why PHP is important:
        https://www.jobsity.com/blog/8-reasons-why-php-is-still-so-important-for-web-development#:~:text=PHP%20(Hypertext%20Preprocessor)%20is%20known,call%20external%20files%20for%20data.

        Project Ideas:
            - Build a Clothes Recommendation System
            - Customer Relationship Management for ISP
            - Predict Movie Success through Data Mining""")

        await ctx.send(response)
    @commands.command()
    async def Javascript (self, ctx):
        response = """
        Here is some resources that I found: 

        Learning Resrouces you might want to start: 
        https://www.w3schools.com/js/

        Another tutorial for Javascript:
        https://www.tutorialspoint.com/javascript/index.htm

        Why KavaScript is important:
        https://www.bigcommerce.com/ecommerce-answers/what-javascript-and-why-it-important/#:~:text=js%20makes%20responsive%20design%20easier,not%20be%20possible%20without%20it.
        Project Ideas:
            - A Simple To-Do List App
            - Create a Simple Timer
            - Build an Image Carousel From Scratch
        """

        await ctx.send(response)


def setup(bot):
    bot.add_cog(main_funcs(bot))  